from openpyxl import load_workbook

ss = load_workbook("workbooktest.xlsx")
sheet = ss.active

categories = ['food', "housing", "entertainment", "supplies", "other"]

sheet.title = "Budget Test"
# sheet['A1':'C2'] = ['Expense Name', 'Category', 'Description', 'Amount']
sheet['A1'] = 'Expense Name'
sheet['B1'] = 'Category'
sheet['C1'] = 'Description'
sheet['D1'] = 'Amount'
maxrow = (sheet.max_row + 1)
nextrow = str(maxrow)


def exit_protocol():
    # sheet['D2:D100'].number_format = '$#,##0_'
    ss.save("workbooktest.xlsx")
    print("Goodbye and Fuck Off")
    # quit()


def ask_me():
    valid_active_response = False
    while not valid_active_response:
        program_active = input("Would you like to enter an expense?(y/n)\n").lower()
        if program_active not in ['y', 'n']:
            print("please put in a valid option")
            valid_active_response = False
        else:
            valid_active_response = True
            break
    if program_active == 'y':
        new_expense()
    else:
        exit_protocol()


def new_expense():
    sheet['A' + nextrow] = input('Enter Expense Name: ')
    category: str = ''
    ok_category = False
    while not ok_category:
        category = input("Enter Category: ").lower()
        if category not in categories:
            print("Please choose one of the following options (Housing, Entertainment, Food, Supplies)")
            ok_category = False
        else:
            ok_category = True
    sheet['B' + nextrow] = str(category)
    sheet['C' + nextrow] = input("Enter Description: ")

    ok_amount = False
    while not ok_amount:
        money = input('Enter Amount: ')
        try:
            money = float(money)
            ok_amount = True
            sheet['D' + nextrow] = money
        except ValueError:
            try:
                money = int(money)
                ok_amount = True
                sheet['D' + nextrow] = '$' + str(money)
            except ValueError:
                print("Don't be a dickhead!")
                ok_amount = False
    ask_me()


ask_me()
# program_active = input("Would you like to enter another expense?(y/n)\n").lower()
# if program_active not in ['y', 'n']:
#     print('please put in a valid option')
#     valid_active_response = False
# else:
#     valid_active_response = True
# if program_active == 'y':
#     new_expense()
# else:
#     exit_protocol()

# valid_active_response = False
# while not valid_active_response:
#     program_active = str.lower(input("""Would you like to enter an expense?(y/n)\n"""))
#     if program_active not in ['y', 'n']:
#         print("Please choose y or n")
#         valid_active_response = False
#     else:
#         valid_active_response = True
# while program_active is "y":
#     new_expense()

# if __name__ == "__main__":